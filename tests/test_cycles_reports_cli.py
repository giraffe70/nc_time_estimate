from __future__ import annotations

import json
import math
from datetime import datetime
import xml.etree.ElementTree as ET
import zipfile

from nc_time_twin import estimate_nc_time, estimate_nc_time_with_comparison
from nc_time_twin.cli import main
from nc_time_twin.core.report import exporter_excel
from nc_time_twin.core.report.auto_outputs import (
    manual_export_path,
    manual_export_path_in_dir,
    write_auto_log,
    write_auto_outputs,
)
from nc_time_twin.core.report.exporters import export_result


def test_event_times_and_dwell(write_nc, profile_path) -> None:
    nc = write_nc(
        """
        G21 G90
        T1 M06
        S1000 M03
        M08
        G04 P1000
        M09
        M05
        """
    )
    result = estimate_nc_time(nc, profile_path)
    assert math.isclose(result.tool_change_time_sec, 8.0, rel_tol=1e-9)
    assert math.isclose(result.spindle_time_sec, 3.0, rel_tol=1e-9)
    assert math.isclose(result.coolant_time_sec, 1.0, rel_tol=1e-9)
    assert math.isclose(result.dwell_time_sec, 1.0, rel_tol=1e-9)


def test_g81_g82_g83_cycles_expand(write_nc, profile_path) -> None:
    nc = write_nc(
        """
        G21 G90 G17
        G81 X20 Y0 Z-5 R2 F300
        G82 X30 Y0 Z-4 R2 P500 F300
        G83 X40 Y0 Z-9 R2 Q3 F300
        G80
        """
    )
    result = estimate_nc_time(nc, profile_path)
    types = [block.display_type for block in result.ir_program]
    assert types[:4] == ["rapid", "rapid", "linear", "rapid"]
    assert "dwell" in types
    assert types.count("linear") >= 5
    assert result.ir_program[-1].end.as_tuple() == (40.0, 0.0, 2.0)


def test_json_csv_excel_html_exports(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path)
    for suffix in ("json", "csv", "xlsx", "html"):
        out = artifact_path(suffix)
        export_result(result, out)
        assert out.exists()
        assert out.stat().st_size > 0


def test_excel_export_uses_consolidated_diagnostics_sheet(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path)
    out = artifact_path("xlsx")

    export_result(result, out)

    sheet_names = _xlsx_sheet_names(out)
    assert "summary" in sheet_names
    assert "blocks" in sheet_names
    assert "diagnostics" in sheet_names
    assert not any(name.startswith("comparison") for name in sheet_names)
    assert not {
        "warnings",
        "feed_histogram",
        "top_slow_blocks",
        "feed_sanity_summary",
        "feed_sanity_issues",
        "feed_recommendation",
    }.intersection(sheet_names)
    assert _xlsx_sheet_first_row(out, "summary") == ["metric", "value"]


def test_comparison_exports_include_segment_differences(write_nc, profile_path, artifact_path) -> None:
    source = write_nc("G21 G90\nG01 X100 F6000")
    candidate = write_nc("G21 G90\nG01 X100 F100")
    result = estimate_nc_time_with_comparison(candidate, source, profile_path)

    xlsx = artifact_path("xlsx")
    export_result(result, xlsx, "xlsx")
    assert "comparison_segments" in _xlsx_sheet_names(xlsx)
    header = _xlsx_sheet_first_row(xlsx, "comparison_segments")
    assert "line_no" in header
    assert "原始 F" in header
    assert "優化後 F" in header
    assert "原始有效 feed" in header
    assert "優化後有效 feed" in header
    assert "原始時間" in header
    assert "優化後時間" in header
    assert "時間差" in header
    assert "是否低速異常" in header
    assert "是否單位疑似異常" in header

    json_out = artifact_path("json")
    export_result(result, json_out, "json")
    data = json.loads(json_out.read_text(encoding="utf-8"))
    assert data["comparison"]["segment_differences"][0]["line_no"] == 2

    csv_out = artifact_path("csv")
    export_result(result, csv_out, "csv")
    csv_text = csv_out.read_text(encoding="utf-8-sig")
    assert "原始 F" in csv_text
    assert "優化後 F" in csv_text

    html_out = artifact_path("html")
    export_result(result, html_out, "html")
    html = html_out.read_text(encoding="utf-8")
    assert "原始 NC vs 優化 NC 逐段差異分析" in html
    assert "優化後 F" in html


def test_excel_export_downsamples_large_phase2_dynamic_sheet(write_nc, profile_path, artifact_path, monkeypatch) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path, time_model="phase2")
    result.phase2_dynamic_samples = [
        {
            "time_sec": float(index),
            "velocity_mm_s": float(index % 100),
            "segment_id": index,
            "block_index": 0,
        }
        for index in range(25)
    ]
    out = artifact_path("xlsx")
    monkeypatch.setattr(exporter_excel, "MAX_EXCEL_PHASE2_DYNAMIC_ROWS", 10)
    monkeypatch.setattr(exporter_excel, "MAX_CHART_PHASE2_DYNAMIC_POINTS", 10)

    export_result(result, out)

    from openpyxl import load_workbook

    workbook = load_workbook(out, read_only=True)
    try:
        assert "phase2_dynamic" in workbook.sheetnames
        assert workbook["phase2_dynamic"].max_row == 11
        diagnostics_values = [
            cell
            for row in workbook["diagnostics"].iter_rows(values_only=True)
            for cell in row
            if isinstance(cell, str)
        ]
        assert any("original_samples=25" in value for value in diagnostics_values)
    finally:
        workbook.close()


def test_excel_export_uses_relative_relationship_targets(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path, time_model="phase2")
    out = artifact_path("xlsx")

    export_result(result, out)

    with zipfile.ZipFile(out) as archive:
        for name in archive.namelist():
            if not name.endswith(".rels"):
                continue
            text = archive.read(name).decode("utf-8")
            assert 'Target="/xl/' not in text


def test_cli_estimate_json(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    out = artifact_path("json")
    exit_code = main(["estimate", "--nc", str(nc), "--profile", str(profile_path), "--out", str(out)])
    assert exit_code == 0
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["total_time_sec"] == 6.0


def test_cli_compare_can_fail_on_regression(write_nc, profile_path, artifact_path) -> None:
    source = write_nc("G21 G90\nG01 X100 F6000")
    candidate = write_nc("G21 G90\nG01 X100 F100")
    out = artifact_path("json")

    exit_code = main(
        [
            "estimate",
            "--nc",
            str(candidate),
            "--compare-nc",
            str(source),
            "--profile",
            str(profile_path),
            "--out",
            str(out),
            "--fail-on-regression",
        ]
    )

    assert exit_code == 1
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["comparison"]["geometry_match"] is True
    assert data["comparison"]["is_regression"] is True
    assert data["comparison"]["top_time_regression_blocks"][0]["candidate_line_no"] == 2
    assert data["comparison"]["max_regression_ratio"] == 0.0
    assert data["comparison"]["regression_ratio"] > 0


def test_cli_strict_feed_can_fail_on_sanity_error(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc(
        """
        G21 G90 G94
        G01 X100 F6
        X101 F104
        X102 F1000000
        """
    )
    out = artifact_path("json")

    exit_code = main(
        [
            "estimate",
            "--nc",
            str(nc),
            "--profile",
            str(profile_path),
            "--out",
            str(out),
            "--strict-feed",
            "--fail-on-sanity-error",
        ]
    )

    assert exit_code == 1
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["feed_unit_effective"] == "mm_per_min"
    assert data["summary"]["feed_sanity_critical_count"] > 0
    assert any(issue["code"] == "mixed_feed_scale" for issue in data["feed_sanity_issues"])


def test_cli_normalize_feed_converts_m_per_min_to_mm_per_min(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc(
        """
        G21 G90 G94
        G01 X1 F6
        X2 F104
        (F5 comment)
        G95 F0.2
        G01 X3
        """
    )
    out = artifact_path("nc")

    exit_code = main(
        [
            "normalize-feed",
            "--nc",
            str(nc),
            "--profile",
            str(profile_path),
            "--input-feed-unit",
            "m_per_min",
            "--out",
            str(out),
        ]
    )

    assert exit_code == 0
    text = out.read_text(encoding="utf-8")
    assert "F6000" in text
    assert "F10000" in text
    assert "(F5 comment)" in text
    assert "G95 F0.2" in text


def test_cli_normalize_feed_keeps_mm_per_min_values(write_nc, profile_path, artifact_path) -> None:
    nc = write_nc(
        """
        G21 G90 G94
        G01 X1 F104
        X2 F12000
        """
    )
    out = artifact_path("nc")

    exit_code = main(
        [
            "normalize-feed",
            "--nc",
            str(nc),
            "--profile",
            str(profile_path),
            "--input-feed-unit",
            "mm_per_min",
            "--out",
            str(out),
        ]
    )

    assert exit_code == 0
    text = out.read_text(encoding="utf-8")
    assert "F104" in text
    assert "F10000" in text


def test_auto_outputs_include_report_chart_data_and_log(write_nc, profile_path, tmp_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path)
    paths = write_auto_outputs(
        result,
        nc,
        base_dir=tmp_path,
        now=datetime(2026, 4, 27, 16, 25),
    )

    assert paths.report_path == tmp_path / "output" / f"Report_{nc.stem}_20260427_1625.xlsx"
    assert paths.log_path == tmp_path / "logs" / f"{nc.stem}_20260427_1625.log"
    assert paths.report_path.exists()
    assert paths.report_path.stat().st_size > 0
    assert "Warnings:" in paths.log_path.read_text(encoding="utf-8")


def test_auto_log_does_not_write_report(write_nc, profile_path, tmp_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path)

    log_path = write_auto_log(
        result,
        nc,
        base_dir=tmp_path,
        now=datetime(2026, 4, 27, 16, 25),
    )

    assert log_path == tmp_path / "logs" / f"{nc.stem}_20260427_1625.log"
    assert log_path.exists()
    assert "Summary:" in log_path.read_text(encoding="utf-8")
    assert not (tmp_path / "output").exists()


def test_manual_export_path_uses_output_directory_and_nc_filename(write_nc, tmp_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    path = manual_export_path(
        nc,
        "json",
        base_dir=tmp_path,
        now=datetime(2026, 4, 27, 16, 25),
    )
    assert path == tmp_path / "output" / f"{nc.stem}_20260427_1625.json"


def test_manual_export_path_in_dir_allows_multiple_formats(write_nc, profile_path, tmp_path) -> None:
    nc = write_nc("G21 G90\nG01 X100 F1000")
    result = estimate_nc_time(nc, profile_path)
    output_dir = tmp_path / "chosen_reports"

    for fmt in ("json", "csv", "xlsx", "html"):
        path = manual_export_path_in_dir(
            nc,
            fmt,
            output_dir,
            now=datetime(2026, 4, 27, 16, 25),
        )
        export_result(result, path, fmt)
        assert path == output_dir / f"{nc.stem}_20260427_1625.{fmt}"
        assert path.exists()
        assert path.stat().st_size > 0


def _xlsx_sheet_names(path) -> list[str]:
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [sheet.attrib["name"] for sheet in workbook.findall(".//main:sheet", namespace)]


def _xlsx_sheet_first_row(path, sheet_name: str) -> list[str]:
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        namespace = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        sheet = next(
            item for item in workbook.findall(".//main:sheet", namespace) if item.attrib["name"] == sheet_name
        )
        rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = next(item.attrib["Target"] for item in rels if item.attrib["Id"] == rel_id)
        worksheet_path = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
        worksheet = ET.fromstring(archive.read(worksheet_path))
    row = worksheet.find(".//main:row", namespace)
    if row is None:
        return []
    values: list[str] = []
    for cell in row.findall("main:c", namespace):
        text = cell.find(".//main:t", namespace)
        values.append(text.text if text is not None and text.text is not None else "")
    return values
